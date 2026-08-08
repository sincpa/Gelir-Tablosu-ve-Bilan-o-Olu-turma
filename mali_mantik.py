"""
mali_mantik.py
--------------
Mizandan Gelir Tablosu ve Bilanco olusturma mantigi.

CALISMA MANTIGI (3 ADIM):
  ADIM 1 - OKUMA VE TESHIS : Mizanlar okunur, 7/A hesaplarinin durumu tespit edilir
  ADIM 2 - 7/A AKTARIMI    : Yansitilmamis kalan tutarlar 6'li hesaplara aktarilir
  ADIM 3 - TABLO YAZIMI    : Hesaplar tabloya yazilir, donem sonucu 590/591'e tasinir

TEMEL KURAL (7/A):
  Kalan tutar = 7'li borc bakiyesi - ilgili yansitma hesabinin alacak bakiyesi

  Bu tek formul uc senaryoyu da dogru cozer:
    - Yansitma TAM yapilmis   -> kalan 0 -> aktarim yok (mukerrer kayit olmaz)
    - Yansitma HIC yapilmamis -> kalan = tamami -> tamami aktarilir
    - Yansitma KISMI yapilmis -> kalan = fark -> sadece fark aktarilir
"""

from dataclasses import dataclass, field

import openpyxl


# ======================================================================
# 7/A MALIYET HESAPLARI -> 6'LI HESAP ESLEMESI
# ======================================================================
# Format:  7'li hesap : (yansitma hesabi, hedef 6'li hesap, aciklama)
YEDILI_ESLEME = {
    "710": ("711", "620", "Direkt Ilk Madde ve Malzeme Giderleri"),
    "720": ("721", "620", "Direkt Iscilik Giderleri"),
    "730": ("731", "620", "Genel Uretim Giderleri"),
    "740": ("741", "622", "Hizmet Uretim Maliyeti"),
    "750": ("751", "630", "Arastirma ve Gelistirme Giderleri"),
    "760": ("761", "631", "Pazarlama, Satis ve Dagitim Giderleri"),
    "770": ("771", "632", "Genel Yonetim Giderleri"),
    "780": ("781", "660", "Finansman Giderleri"),
}

# Tabloya ASLA dogrudan yazilmayacak hesap gruplari
# 7'liler aktarim yoluyla islenir, 69'lar donem sonu kapanis hesaplaridir
HARIC_ONEKLER = ("7", "69")


# ======================================================================
# VERI YAPILARI
# ======================================================================
def turkce_sadelestir(metin: str) -> str:
    """
    Turkce karakterleri sadelestirir ve kucuk harfe cevirir.
    'Borç Bakiye' -> 'borc bakiye'   (baslik eslestirmesi icin)
    """
    if not isinstance(metin, str):
        return ""
    esleme = {
        "ç": "c", "Ç": "c", "ğ": "g", "Ğ": "g", "ı": "i", "I": "i",
        "İ": "i", "i": "i", "ö": "o", "Ö": "o", "ş": "s", "Ş": "s",
        "ü": "u", "Ü": "u",
    }
    return "".join(esleme.get(ch, ch) for ch in metin).lower().strip()


@dataclass
class MizanSatiri:
    kod: str
    ad: str
    borc_bakiye: float = 0.0
    alacak_bakiye: float = 0.0

    @property
    def net(self) -> float:
        """Isaretli net bakiye: pozitif = borc, negatif = alacak"""
        return self.borc_bakiye - self.alacak_bakiye

    @property
    def tutar(self) -> float:
        """
        Tablolara yazilacak MUTLAK tutar.

        Neden mutlak deger? Cunku sablonun kendi formulleri isareti zaten
        yonetiyor. Ornegin '103 Verilen Cekler (-)' hesabi sablonda
        =SUM(D10,D11,D12,-D13,D14) seklinde CIKARILIYOR. Biz oraya eksi
        deger yazarsak cift olumsuzlama olur ve tutar yanlis cikar.
        """
        return abs(self.net)


@dataclass
class Teshis:
    """ADIM 1 ciktisi: 7/A hesaplarinin durum raporu"""
    kod: str
    ad: str
    yedili_bakiye: float
    yansitma_bakiye: float
    kalan: float
    hedef: str
    durum: str  # "Yansitma tam" | "Yansitma yok" | "Kismi yansitma"


@dataclass
class Sonuc:
    teshisler: list = field(default_factory=list)
    aktarimlar: list = field(default_factory=list)
    uyarilar: list = field(default_factory=list)
    donem_sonucu: float = 0.0
    aktif_toplam: float = 0.0
    pasif_toplam: float = 0.0
    yazilan_hucre: int = 0
    eslesmeyen: list = field(default_factory=list)


# ======================================================================
# ADIM 1 - MIZAN OKUMA
# ======================================================================
def mizan_oku(ws) -> dict:
    """
    Mizan sayfasini okur ve {hesap_kodu: MizanSatiri} sozlugu dondurur.

    Beklenen sutun duzeni (baslik satiri otomatik bulunur):
      Hesap Kodu | Hesap Adi | Borc Tutari | Alacak Tutari | Borc Bakiye | Alacak Bakiye

    Not: Borc/Alacak TUTARI degil, BAKIYESI kullanilir.
    """
    baslik_satiri = None
    sutunlar = {}

    # Baslik satirini bul (ilk 30 satirda "hesap kodu" gecen satir)
    for r in range(1, min(ws.max_row, 30) + 1):
        hucreler = {}
        for c in range(1, min(ws.max_column, 15) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                hucreler[c] = turkce_sadelestir(v)
        if not hucreler:
            continue
        metin = " ".join(hucreler.values())
        if "hesap kodu" in metin or ("hesap" in metin and "bakiye" in metin):
            baslik_satiri = r
            for c, v in hucreler.items():
                if "hesap kodu" in v or v in ("kod", "hesap no"):
                    sutunlar["kod"] = c
                elif "hesap adi" in v or v in ("ad", "aciklama", "hesap ismi"):
                    sutunlar["ad"] = c
                elif "borc" in v and "bakiye" in v:
                    sutunlar["borc_bakiye"] = c
                elif "alacak" in v and "bakiye" in v:
                    sutunlar["alacak_bakiye"] = c
            break

    if baslik_satiri is None or "kod" not in sutunlar:
        raise ValueError(
            "Mizan sayfasinda baslik satiri bulunamadi. "
            "Sayfada 'Hesap Kodu' ve 'Borc Bakiye' / 'Alacak Bakiye' "
            "basliklarinin bulundugundan emin olun."
        )

    if "borc_bakiye" not in sutunlar or "alacak_bakiye" not in sutunlar:
        raise ValueError(
            "Mizanda 'Borc Bakiye' ve/veya 'Alacak Bakiye' sutunu bulunamadi."
        )

    mizan = {}
    for r in range(baslik_satiri + 1, ws.max_row + 1):
        ham_kod = ws.cell(row=r, column=sutunlar["kod"]).value
        if ham_kod is None:
            continue

        kod = str(ham_kod).strip()
        # Sayisal olmayan satirlari atla (GENEL TOPLAM, notlar vb.)
        if not kod.isdigit():
            continue
        # Alt kirilimli hesaplari ana hesaba indirge (600.01.001 -> 600)
        kod = kod.split(".")[0].strip()
        if len(kod) < 3:
            continue
        ana_kod = kod[:3]

        ad = ""
        if "ad" in sutunlar:
            ham_ad = ws.cell(row=r, column=sutunlar["ad"]).value
            ad = str(ham_ad).strip() if ham_ad else ""

        def sayi(sutun_adi):
            v = ws.cell(row=r, column=sutunlar[sutun_adi]).value
            if v is None:
                return 0.0
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        borc = sayi("borc_bakiye")
        alacak = sayi("alacak_bakiye")

        if ana_kod in mizan:
            mizan[ana_kod].borc_bakiye += borc
            mizan[ana_kod].alacak_bakiye += alacak
        else:
            mizan[ana_kod] = MizanSatiri(ana_kod, ad, borc, alacak)

    return mizan


# ======================================================================
# ADIM 2 - 7/A AKTARIMI
# ======================================================================
def yedili_aktarim(mizan: dict, sonuc: Sonuc):
    """
    7/A maliyet hesaplarindaki yansitilmamis kalan tutarlari 6'li
    hesaplara aktarir.

    Kalan = 7'li borc bakiyesi - yansitma hesabinin alacak bakiyesi
    """
    for yedili, (yansitma, hedef, aciklama) in YEDILI_ESLEME.items():
        y_satir = mizan.get(yedili)
        if y_satir is None:
            continue

        yedili_bakiye = y_satir.borc_bakiye
        if abs(yedili_bakiye) < 0.005:
            continue

        yans_satir = mizan.get(yansitma)
        yansitma_bakiye = yans_satir.alacak_bakiye if yans_satir else 0.0

        kalan = round(yedili_bakiye - yansitma_bakiye, 2)

        # Durum teshisi
        if abs(kalan) < 0.005:
            durum = "Yansitma tam"
        elif abs(yansitma_bakiye) < 0.005:
            durum = "Yansitma yok"
        else:
            durum = "Kismi yansitma"

        sonuc.teshisler.append(
            Teshis(
                kod=yedili,
                ad=aciklama,
                yedili_bakiye=yedili_bakiye,
                yansitma_bakiye=yansitma_bakiye,
                kalan=kalan,
                hedef=hedef,
                durum=durum,
            )
        )

        if kalan < -0.005:
            sonuc.uyarilar.append(
                f"{yedili} hesabinda yansitma ({yansitma_bakiye:,.2f}) "
                f"gider bakiyesinden ({yedili_bakiye:,.2f}) BUYUK. "
                f"Aktarim yapilmadi, mizani kontrol edin."
            )
            continue

        if kalan < 0.005:
            continue  # aktarilacak bir sey yok

        # Kalani hedef 6'li hesaba BORC olarak ekle
        if hedef in mizan:
            mizan[hedef].borc_bakiye += kalan
        else:
            mizan[hedef] = MizanSatiri(hedef, f"({yedili} aktarimi)", kalan, 0.0)

        sonuc.aktarimlar.append(
            {
                "kaynak": yedili,
                "hedef": hedef,
                "tutar": kalan,
                "aciklama": aciklama,
                "durum": durum,
            }
        )


# ======================================================================
# DONEM SONUCU
# ======================================================================
def donem_sonucu_hesapla(mizan: dict) -> float:
    """
    Donem kari/zarari = 6'li hesaplarin (alacak - borc) toplami

    Pozitif -> KAR   (590'a yazilir)
    Negatif -> ZARAR (591'e yazilir)

    NOT: Vergi karsiligi HESAPLANMAZ. Bu geciciv vergi donemi tablolaridir;
    vergi hesabi ayri sayfada elle yapilir.
    """
    toplam = 0.0
    for kod, satir in mizan.items():
        if not kod.startswith("6"):
            continue
        if kod.startswith("69"):  # kapanis hesaplari haric
            continue
        toplam += satir.alacak_bakiye - satir.borc_bakiye
    return round(toplam, 2)


def aktif_pasif_hesapla(mizan: dict, donem_sonucu: float):
    """Bagimsiz Aktif = Pasif kontrolu icin toplamlari hesaplar."""
    aktif = 0.0
    pasif = 0.0
    for kod, satir in mizan.items():
        if kod[0] in ("1", "2"):
            aktif += satir.net              # borc - alacak
        elif kod[0] in ("3", "4", "5"):
            if kod.startswith("59"):
                continue                    # donem sonucunu biz hesapliyoruz
            pasif += -satir.net             # alacak - borc
    pasif += donem_sonucu
    return round(aktif, 2), round(pasif, 2)


# ======================================================================
# SABLON HARITALARI
# ======================================================================
def gelir_tablosu_haritasi(ws):
    """
    Gelir Tablosu sayfasindan {hesap_kodu: satir_no} haritasi cikarir.

    KURAL: A sutununda hesap kodu VARSA -> veri satiri (program yazar)
           A sutunu BOS ise             -> ara toplam formulu (dokunulmaz)
    """
    harita = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        kod = str(v).strip()
        if kod.isdigit() and len(kod) == 3:
            harita[kod] = r
    return harita


def bilanco_haritasi(ws):
    """
    Bilanco sayfasindan iki blok icin harita cikarir.

    Sol blok (AKTIF) : kod B sutununda, tutarlar D (onceki) ve E (cari)
    Sag blok (PASIF) : kod F sutununda, tutarlar H (onceki) ve I (cari)

    KURAL: Sadece 3 HANELI kodlar veri satiridir.
           1-2 haneli kodlar grup toplamidir (formul) -> dokunulmaz.
    """
    aktif, pasif = {}, {}
    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if b is not None:
            kod = str(b).strip()
            if kod.isdigit() and len(kod) == 3:
                aktif[kod] = r
        f = ws.cell(row=r, column=6).value
        if f is not None:
            kod = str(f).strip()
            if kod.isdigit() and len(kod) == 3:
                pasif[kod] = r
    return aktif, pasif


def alt_kirilim_satirlari(ws):
    """
    Gelir tablosundaki '7-a-', '7-b-', '4-a-', '4-b-' gibi alt kirilim
    satirlarini bulur. Bunlarin hesap kodu YOKTUR ama eski veri tasiyabilir.
    Temizlenmeleri gerekir, yoksa onceki mukellefin verisi ekranda kalir.
    """
    import re as _re
    satirlar = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and _re.search(r"\d+\s*-\s*[abcçd]\s*-", v):
            satirlar.append(r)
    return satirlar
