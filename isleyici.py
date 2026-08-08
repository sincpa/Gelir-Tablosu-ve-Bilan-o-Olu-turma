"""
isleyici.py
-----------
Ana islem akisi. Mizanlari okur, 7/A aktarimini yapar, tablolari doldurur
ve mizan sayfalarini silerek dosyayi geri verir.

SABIT KURALLAR:
  * Sayfa adlari TAM olarak 'mizan1' ve 'mizan2' olmalidir (buyuk/kucuk harf duyarli)
  * mizan1 -> onceki donem (gecen yil)
  * mizan2 -> cari donem   (bu yil)
  * mizan2 yoksa program CALISMAZ
  * Yuklenmeyen donemin sutunu TEMIZLENIR (bos birakilir)
  * Hucrelere FORMUL degil SABIT SAYI yazilir
    -> boylece mizan sayfalari silinince tablo bozulmaz
  * Mizan sayfalari ciktidan SILINIR (dosya dogrudan mukellefe gidebilir)
"""

from io import BytesIO

import openpyxl

from mali_mantik import (
    Sonuc,
    aktif_pasif_hesapla,
    bilanco_haritasi,
    donem_sonucu_hesapla,
    gelir_tablosu_haritasi,
    mizan_oku,
    yedili_aktarim,
)
from xlsx_cerrahi import ExcelCerrah

# Sayfa adlari (TAM eslesme sarti)
MIZAN_ONCEKI = "mizan1"
MIZAN_CARI = "mizan2"

GELIR_TABLOSU_SAYFA = "GELİR TABLOSU"
BILANCO_SAYFA = "BİLANÇO"

# Sutun haritasi
GT_SUTUN = {"onceki": "C", "cari": "E"}          # Gelir Tablosu
BIL_AKTIF_SUTUN = {"onceki": "D", "cari": "E"}   # Bilanco sol blok
BIL_PASIF_SUTUN = {"onceki": "H", "cari": "I"}   # Bilanco sag blok


def _sayfa_bul(wb_openpyxl, aday_adlar):
    """Sablon sayfa adini bulur (bosluk/harf farklarina toleransli)."""
    for ad in wb_openpyxl.sheetnames:
        if ad in aday_adlar:
            return ad
    # Toleransli arama
    normal = lambda s: s.strip().upper().replace("İ", "I").replace("Ç", "C")
    hedefler = {normal(a) for a in aday_adlar}
    for ad in wb_openpyxl.sheetnames:
        if normal(ad) in hedefler:
            return ad
    return None


def isle(dosya_baytlari: bytes):
    """
    Ana islem fonksiyonu.

    Girdi : Sablon Excel dosyasinin baytlari (icinde mizan1 / mizan2 sayfalari)
    Cikti : (yeni_dosya_baytlari, Sonuc)
    """
    sonuc = Sonuc()

    # ------------------------------------------------------------------
    # Dosyayi iki sekilde aciyoruz:
    #   openpyxl -> okumak icin (hucre degerlerini kolay okur)
    #   cerrah   -> yazmak icin (bicimlendirmeyi bozmaz)
    # ------------------------------------------------------------------
    wb = openpyxl.load_workbook(BytesIO(dosya_baytlari), data_only=True)
    cerrah = ExcelCerrah(dosya_baytlari)

    # ------------------------------------------------------------------
    # GUVENLIK KONTROLU: sayfa adlari TAM eslesmeli
    # ------------------------------------------------------------------
    var_onceki = MIZAN_ONCEKI in wb.sheetnames
    var_cari = MIZAN_CARI in wb.sheetnames

    if not var_cari:
        # Yanlis yazilmis olabilecek adlari tespit edip aciklayici hata verelim
        benzer = [
            ad for ad in wb.sheetnames
            if ad.strip().lower().replace(" ", "").startswith("mizan")
            and ad not in (MIZAN_ONCEKI, MIZAN_CARI)
        ]
        mesaj = (
            f"'{MIZAN_CARI}' adinda bir sayfa bulunamadi. "
            f"Program calismadi, dosyaya dokunulmadi."
        )
        if benzer:
            mesaj += (
                f"\n\nDosyada su sayfalar var: {', '.join(repr(b) for b in benzer)}. "
                f"Sayfa adi TAM olarak 'mizan1' ve 'mizan2' olmalidir "
                f"(kucuk harf, bosluksuz)."
            )
        raise ValueError(mesaj)

    # ------------------------------------------------------------------
    # Sablon sayfalarini bul
    # ------------------------------------------------------------------
    gt_ad = _sayfa_bul(wb, {GELIR_TABLOSU_SAYFA})
    bil_ad = _sayfa_bul(wb, {BILANCO_SAYFA})
    if gt_ad is None:
        raise ValueError("Sablonda 'GELİR TABLOSU' sayfasi bulunamadi.")
    if bil_ad is None:
        raise ValueError("Sablonda 'BİLANÇO' sayfasi bulunamadi.")

    gt_ws = wb[gt_ad]
    bil_ws = wb[bil_ad]

    # ------------------------------------------------------------------
    # ADIM 1 - MIZANLARI OKU
    # ------------------------------------------------------------------
    mizanlar = {}
    if var_onceki:
        mizanlar["onceki"] = mizan_oku(wb[MIZAN_ONCEKI])
    if var_cari:
        mizanlar["cari"] = mizan_oku(wb[MIZAN_CARI])

    if not var_onceki:
        sonuc.uyarilar.append(
            "'mizan1' sayfasi yok. Onceki donem sutunlari TEMIZLENDI (bos birakildi)."
        )

    # ------------------------------------------------------------------
    # ADIM 2 - 7/A AKTARIMI + DONEM SONUCU
    # ------------------------------------------------------------------
    donem_sonuclari = {}
    for donem, mizan in mizanlar.items():
        # Teshis ve aktarimlari donem etiketiyle isaretleyelim
        onceki_teshis_sayisi = len(sonuc.teshisler)
        onceki_aktarim_sayisi = len(sonuc.aktarimlar)

        yedili_aktarim(mizan, sonuc)

        etiket = "Onceki donem (mizan1)" if donem == "onceki" else "Cari donem (mizan2)"
        for t in sonuc.teshisler[onceki_teshis_sayisi:]:
            t.ad = f"{t.ad}"
            setattr(t, "donem", etiket)
        for a in sonuc.aktarimlar[onceki_aktarim_sayisi:]:
            a["donem"] = etiket

        ds = donem_sonucu_hesapla(mizan)
        donem_sonuclari[donem] = ds

        aktif, pasif = aktif_pasif_hesapla(mizan, ds)
        if donem == "cari":
            sonuc.donem_sonucu = ds
            sonuc.aktif_toplam = aktif
            sonuc.pasif_toplam = pasif

        fark = round(aktif - pasif, 2)
        if abs(fark) > 0.5:
            sonuc.uyarilar.append(
                f"{etiket}: Aktif ({aktif:,.2f}) ile Pasif ({pasif:,.2f}) "
                f"esit degil. Fark: {fark:,.2f}. Mizani kontrol edin."
            )

        # Mizanda zaten 590/591 varsa kullaniciyi bilgilendir
        for k in ("590", "591"):
            if k in mizan and abs(mizan[k].net) > 0.005:
                sonuc.uyarilar.append(
                    f"{etiket}: Mizanda {k} hesabi bakiyeli. Tabloya mizandaki deger "
                    f"degil, 6'li hesaplardan HESAPLANAN donem sonucu yazildi."
                )

    # ------------------------------------------------------------------
    # ADIM 3 - TABLOLARI TEMIZLE VE DOLDUR
    # ------------------------------------------------------------------
    gt_harita = gelir_tablosu_haritasi(gt_ws)
    bil_aktif_harita, bil_pasif_harita = bilanco_haritasi(bil_ws)
    # (alt kirilim satirlari artik formul-farkindali temizlik ile kapsaniyor)

    # --- 3a. TEMIZLIK ---------------------------------------------------
    # Her calistirmada veri bolgesindeki TUM sayi hucreleri bosaltilir.
    #
    # NEDEN HESAP KODUNA GORE DEGIL?
    # Sablonda bazi hesaplarin rakami, hesap kodunun BIR ALT satirinda durur
    # (hesap adi iki satira bolundugu icin). Ornek: 370 hesabinin kodu 47.
    # satirda, rakami 48. satirda. Sadece kodlu satirlari temizleseydik, o
    # rakamlar onceki mukelleften kalir ve tablo sessizce yanlis cikardi.
    #
    # Bu yuzden kural: veri bolgesindeki FORMUL ICERMEYEN her hucre temizlenir.
    # Formul iceren hucrelere (ara toplamlar, dikey/yatay %) ASLA dokunulmaz.
    wb_formul = openpyxl.load_workbook(BytesIO(dosya_baytlari), data_only=False)

    def veri_bolgesi_temizle(sayfa_ad, ws_f, sutun_harfi, bas_satir, son_satir):
        sutun_no = openpyxl.utils.column_index_from_string(sutun_harfi)
        for r in range(bas_satir, son_satir + 1):
            v = ws_f.cell(row=r, column=sutun_no).value
            if isinstance(v, str) and v.startswith("="):
                continue  # formul -> korunur
            cerrah.hucre_yaz(sayfa_ad, f"{sutun_harfi}{r}", None)

    gt_f = wb_formul[gt_ad]
    bil_f = wb_formul[bil_ad]

    # Veri bolgesi sinirlari: ilk ve son hesap kodu satiri arasi
    if gt_harita:
        gt_bas, gt_son = min(gt_harita.values()), max(gt_harita.values())
        for donem in ("onceki", "cari"):
            veri_bolgesi_temizle(gt_ad, gt_f, GT_SUTUN[donem], gt_bas, gt_son)

    if bil_aktif_harita:
        a_bas, a_son = min(bil_aktif_harita.values()), max(bil_aktif_harita.values())
        for donem in ("onceki", "cari"):
            veri_bolgesi_temizle(bil_ad, bil_f, BIL_AKTIF_SUTUN[donem], a_bas, a_son)

    if bil_pasif_harita:
        p_bas, p_son = min(bil_pasif_harita.values()), max(bil_pasif_harita.values())
        for donem in ("onceki", "cari"):
            veri_bolgesi_temizle(bil_ad, bil_f, BIL_PASIF_SUTUN[donem], p_bas, p_son)

    # --- 3b. DOLDURMA ---------------------------------------------------
    for donem, mizan in mizanlar.items():
        gt_sutun = GT_SUTUN[donem]
        bil_a_sutun = BIL_AKTIF_SUTUN[donem]
        bil_p_sutun = BIL_PASIF_SUTUN[donem]

        yazilan_kodlar = set()

        for kod, satir in sorted(mizan.items()):
            # 7'li ve 69'lu hesaplar tabloya dogrudan yazilmaz
            if kod.startswith("7") or kod.startswith("69"):
                continue
            # Donem sonucunu asagida ayrica yaziyoruz
            if kod in ("590", "591"):
                continue

            tutar = satir.tutar
            if abs(tutar) < 0.005:
                continue  # sifir bakiyeli hesabi bos birak (temiz gorunum)

            hedef_bulundu = False

            if kod in gt_harita:
                cerrah.hucre_yaz(gt_ad, f"{gt_sutun}{gt_harita[kod]}", tutar)
                sonuc.yazilan_hucre += 1
                hedef_bulundu = True

            if kod in bil_aktif_harita:
                cerrah.hucre_yaz(bil_ad, f"{bil_a_sutun}{bil_aktif_harita[kod]}", tutar)
                sonuc.yazilan_hucre += 1
                hedef_bulundu = True

            if kod in bil_pasif_harita:
                cerrah.hucre_yaz(bil_ad, f"{bil_p_sutun}{bil_pasif_harita[kod]}", tutar)
                sonuc.yazilan_hucre += 1
                hedef_bulundu = True

            if hedef_bulundu:
                yazilan_kodlar.add(kod)
            else:
                etiket = "mizan1" if donem == "onceki" else "mizan2"
                sonuc.eslesmeyen.append(
                    {
                        "donem": etiket,
                        "kod": kod,
                        "ad": satir.ad,
                        "tutar": tutar,
                    }
                )

        # --- Donem sonucunu 590 / 591'e yaz ---
        ds = donem_sonuclari[donem]
        if ds >= 0:
            if "590" in bil_pasif_harita:
                cerrah.hucre_yaz(
                    bil_ad, f"{bil_p_sutun}{bil_pasif_harita['590']}", ds if ds else None
                )
                sonuc.yazilan_hucre += 1
            if "591" in bil_pasif_harita:
                cerrah.hucre_yaz(bil_ad, f"{bil_p_sutun}{bil_pasif_harita['591']}", None)
        else:
            if "591" in bil_pasif_harita:
                cerrah.hucre_yaz(
                    bil_ad, f"{bil_p_sutun}{bil_pasif_harita['591']}", abs(ds)
                )
                sonuc.yazilan_hucre += 1
            if "590" in bil_pasif_harita:
                cerrah.hucre_yaz(bil_ad, f"{bil_p_sutun}{bil_pasif_harita['590']}", None)

    # ------------------------------------------------------------------
    # ADIM 4 - MIZAN SAYFALARINI SIL
    # ------------------------------------------------------------------
    # Hucrelere sabit sayi yazdigimiz icin bu silme islemi tabloyu BOZMAZ.
    for ad in (MIZAN_ONCEKI, MIZAN_CARI):
        if cerrah.sayfa_var_mi(ad):
            cerrah.sayfa_sil(ad)

    return cerrah.kaydet_bytes(), sonuc
